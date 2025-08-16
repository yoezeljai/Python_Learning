#Script to ready data from EXCEL
from openpyxl import load_workbook
from openpyxl import Workbook
myexcel=load_workbook(filename=r'/users/yoezeljai/desktop/test.xlsx')
mysheet=myexcel.active
rows=mysheet.max_row
cols=mysheet.max_column
n=[]
a=[]
s=[]
startpos=2

while(rows!=0):
    namestr='A'+str(startpos)
    agestr='B'+str(startpos)
    salstr='C'+str(startpos)
    nameval=mysheet[namestr].value
    ageval=mysheet[agestr].value
    salval=mysheet[salstr].value
    n.append(nameval)
    a.append(ageval)
    s.append(salval)
    rows-=1
    startpos+=1
n2=['x','y','z']
a2=[34,35,36]
s2=[10000,110000,12000]
temp=rows
startpos2=temp+1
rowscount2=len(n2)
listindex=0
while(rowscount2!=0):
    namestr='A'+str(startpos2)
    agestr='B'+str(startpos2)
    salstr='C'+str(startpos2)
    mysheet[namestr]=n[listindex]
    mysheet[agestr]=a[listindex]
    mysheet[salstr]=s[listindex]
    rowscount2-=1
    startpos+=1
    listindex+=1
myexcel.save(r'/users/yoezeljai/desktop/test.xlsx')
print(n)
print(a)
print(s)

