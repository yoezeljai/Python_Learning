#Script for OPENPYXL
from openpyxl import Workbook
wb=Workbook()       #Create a new workbook
ws=wb.active        #makes the wb active
ws1=wb.create_sheet('mysheet') # insert at the end (default)
ws2=wb.create_sheet('mysheet2',0)   # insert at the first position
ws.title='JP'   # change the tile of the sheet
ws.sheet_properties.tabcolor='1072BA'   # changes the sheet tab color
print(wb.sheetnames)    # prints the name of the sheets on workbook
for sheet in wb:
    print(sheet.title)  #prints the list of sheets of Workbook

source=wb.active
target=wb.copy_worksheet(source)

#Accessing one cell
c=ws['A4']
ws['A4']='amazon'

d=ws.cell(row=4,column=2,value='inchuk')

# the below code will create cell of 30*15 grid with values of even number
i=0
for x in range(1,30):
    for y in range(1,15):
        ws.cell(row=x,column=y, value=i)
        i+=2

# Accessing many cells

cell_range=ws1['A1':'C2']
colC=ws1['C']
col_range=ws1['C:D']
row10=ws1[10]
row_range=ws1[5:10]

for row in ws1.iter_rows(min_row=1,max_col=3,max_row=2):
    for cell in row:
        print(cell)

for col in ws1.iter_cols(min_col=1,max_row=3,max_col=2):
    for cell in col:
        print(cell)
print(tuple(ws1.rows))  #displays all the rows in workbook
print(tuple(ws1.columns))  #displays all the columns in workbook

# Accessing values
for row in ws.values:
    print(row)

for columns in ws.values:
    print(columns)

for row in ws.iter_rows(min_row=1,max_col=3,max_row=3):
        print(row)

for cols in ws.iter_cols(min_col=1,max_row=3,max_col=4):
    print(cols)


#Storing data
c.value='hello'   
print(c.value)
d.value='lele'
print(d.value)

# Saving to a file
wb.save('newfile.xlsx')


from openpyxl import load_workbook
wb2=load_workbook('/users/yoezeljai/desktop/test.xlsx')
print(wb2.sheetnames)
wsn=wb2.active
wsn.title='Mysheet'
wsn1=wb2.create_sheet('Raw Data')
print(wb2.sheetnames)

wb2.save('test.xlsx')
