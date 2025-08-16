# Script to load an excel file and
'''from openpyxl import load_workbook
from openpyxl.chart import(AreaChart,Reference,Series,)
wb=load_workbook(r'/users/yoezeljai/desktop/python programs/openpyxl_data/test.xlsx')
print(wb.sheetnames)
ws1=wb.active

chart=AreaChart()
chart.tile='IPH chart'
chart.style=13
chart.x_axis.tile='IPH'
chart.y_axis.tile='LOGIN'

cats=Reference(ws1,min_col=2,min_row=1,max_row=40)
data=Reference(ws1,min_col=2,min_row=1,max_col=5,max_row=40)
chart.add_data(data,titles_from_data=True)
chart.set_categories(cats)
ws1.add_chart(chart,'F10')'''

#Creating Pivot

import pandas as pd
import numpy as np
df=pd.read_excel(r'/users/yoezeljai/desktop/python programs/openpyxl_data/test.xlsx')
df.head()
pd.pivot_table(df,index=['Manager'])

#wb.save(r'/users/yoezeljai/desktop/python programs/openpyxl_data/test.xlsx')
