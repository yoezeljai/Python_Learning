#Script to merge and unmerge cells
from openpyxl import Workbook
from openpyxl.styles import Alignment 
wb=Workbook()
ws=wb.active

ws.merge_cells('A1:D3')     # to merge cell
cell=ws.cell(row=1,column=1,value='INCHUK') # to set value of the cell
cell.alignment=Alignment(horizontal='center',vertical='center') # alignmnet of cell

#Inserting Images
from openpyxl.drawing.image import Image
img=Image(r'/Users/yoezeljai/Desktop/screenshot.png')
ws.add_image(img,'A5')

wb.save(r'/Users/yoezeljai/Desktop/Python Programs/openpyxl_data/test3.xlsx')
