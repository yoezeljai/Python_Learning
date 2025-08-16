from openpyxl import Workbook
wb=Workbook()
ws=wb.active
i=0
for x in range(1,30):
    ws.cell(row=x,column=1,value=i)
    i+=2
j=0
for y in range(1,30):
    ws.cell(row=y,column=2,value=j)
    j+=3
k=0
for z in range(1,30):
    ws.cell(row=z,column=3,value=A[k]+B[k])
    k+=1

                
#ws['C1']='=SUM(A1,B1)'
#ws['D1']='=AVERAGE(A1,B1)'
wb.save(r'/Users/yoezeljai/Desktop/Python Programs/openpyxl_data/test2.xlsx')
